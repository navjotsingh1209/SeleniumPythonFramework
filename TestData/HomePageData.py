import openpyxl

class HomePageData:

    test_HomePage_data = [{"FirstName":"Rahul","LastName":"shetty","Gender":"Male"}, {"FirstName":"Anshika", "LastName":"shetty", "Gender":"Female"}]

    @staticmethod
    def get_test_data(testcasename):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Navjo\\Desktop\\Selenuim project\\Book.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcasename:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]