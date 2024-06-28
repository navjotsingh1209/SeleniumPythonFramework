import openpyxl

Dict = {}
book = openpyxl.load_workbook("C:\\Users\\Navjo\\Desktop\\Selenuim project\\ExcelBook.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
#print(cell.value)

#write io the sheet

#sheet.cell(row=6, column=2).value = "Navpreet"
#sheet.cell(row=5, column=3).value = "Bali"


#print(sheet.max_row)
#print(sheet.max_column)

#print(sheet["B4"].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "tc2":
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)